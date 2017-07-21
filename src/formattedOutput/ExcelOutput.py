'''
Created on 13.07.2017

@author: Jascha Riedel
'''

import openpyxl
from openpyxl.utils import get_column_letter
from evaluation.CellDataObject import CellDataObject
from formattedOutput.ResultsSheet import fillResultsSheet
from formattedOutput.CellDataObjectSheet import createCellDataObjectSheet

SHEET_DATA_TITLE = "DataSheet"
SHEET_RESULTS_TITLE = "Results"
SHEET_DATA_PREFIX = "Data-"


def _createWorkbook(excelFile):
    wb = openpyxl.Workbook()
    wb.active.title = SHEET_RESULTS_TITLE
    
    return wb





def saveCellDataObjects(excelFile, cellDataObjects):
    if isinstance(cellDataObjects, (list, tuple)) is True:
        for dataObject in cellDataObjects:
            if isinstance(dataObject, CellDataObject) is False:
                raise ValueError('Can only process CellDataObjects!')
    else:
        raise ValueError('Parameter cellDataObjects must be list or tuple!')
    
    if isinstance(excelFile, str) is False:
        raise ValueError('Parameter excelFile must be str!')
    
    wb = _createWorkbook(excelFile)
    fillResultsSheet(wb.get_sheet_by_name(SHEET_RESULTS_TITLE), cellDataObjects)
    for idx,cellDataObject in enumerate(cellDataObjects, start = 2):
        sheet = wb.create_sheet(SHEET_DATA_PREFIX + str(cellDataObject.Id), idx)
        createCellDataObjectSheet(sheet, cellDataObject)

    
    wb.save(excelFile)


