'''
Created on 18.07.2017

@author: Jascha Riedel
'''

import numpy as np
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
    )

from openpyxl.utils import column_index_from_string as get_col_index
from openpyxl.chart.axis import NumericAxis


SUMMARAY_HEADINGS_AND_ATTRIBS = (
    ['Area', 'Area'],
    ['Voc', 'Voc'],
    ['Isc', 'Isc'],
    ['Jsc[A/cm^2]', 'Jsc'],
    ['R_p', 'Rp'],
    ['R_s', 'Rs'],
    ['Pmpp[W/cm^2]', 'MppA'],
    ['jsc[A/cm^2]', 'Jsc'],
    ['FF', 'FF'],
    ['Eff[%]', 'Eff']
    )

DATA_HEADINGS = (
    'Voltage',
    'Current',
    'Abs(U)',
    'Abs(I)',
    'j[A/cm^2]',
    'Abs(j)',
    'P[W/cm^2]'
    )

PLOT_TITLES = (
    'U - I Chart',
    'Voltage [V]',
    'Current [I]'
    )

PLOT_SPECIAL_COLUMNS = {
    'Rs': get_col_index('X')
    }


def createCellDataObjectSheet(sheet, cellDataObject):
    
    _createSummaryHeading(sheet, cellDataObject)
    _createDataColumns(sheet, cellDataObject)
    _createDataPlot(sheet, cellDataObject)
    
def _createSummaryHeading(sheet, cellDataObject):
    for idx,headingAndAttribute in enumerate(SUMMARAY_HEADINGS_AND_ATTRIBS, start = 3):
        sheet.cell(row = 1, column = idx).value = headingAndAttribute[0]
        sheet.cell(row = 2, column = idx).value = getattr(cellDataObject, headingAndAttribute[1])
    
    sheet.cell(row = 2, column = 1).value = cellDataObject.Id


def _createDataColumns(sheet, cellDataObject):
    for idx,heading in enumerate(DATA_HEADINGS, start = 1):
        sheet.cell(row = 3, column = idx).value = heading
    
    data = cellDataObject.data
    
    for idx,dataPair in enumerate(data, start = 4):
        sheet.cell(row = idx, column = 1).value = dataPair[0]
        sheet.cell(row = idx, column = 2).value = dataPair[1]
        sheet.cell(row = idx, column = 3).value = '=ABS(A{})'.format(idx)
        sheet.cell(row = idx, column = 4).value = '=ABS(B{})'.format(idx)
        sheet.cell(row = idx, column = 5).value = '=B{}/C2'.format(idx)
        sheet.cell(row = idx, column = 6).value = '=ABS(E{})'.format(idx)
        sheet.cell(row = idx, column = 7).value = '=E{}*A{}'.format(idx,idx)

def _createDataPlot(sheet, cellDataObject):
    chart = ScatterChart()
    chart.title = PLOT_TITLES[0]
    chart.style = 13
    chart.x_axis.title = PLOT_TITLES[1]
    chart.y_axis.title = PLOT_TITLES[2]
    chart.x_axis.crosses = None
    chart.x_axis.crossesAt = 0.0
    chart.y_axis.crosses = None
    chart.y_axis.crossesAt = 0.0
    chart.legend = None
    
    dataX = Reference(sheet, min_col = 1, min_row = 4, max_row = len(cellDataObject.data[:,0]) + 3)
    dataY = Reference(sheet, min_col = 2, min_row = 3, max_row = len(cellDataObject.data[:,0]) + 3)
    
    series = Series(dataY , xvalues = dataX, title_from_data=True)
    chart.append(series)
    chart.append(_createRsSeries(sheet, cellDataObject))
    sheet.add_chart(chart, 'K10')


def _createRsSeries(sheet,cellDataObject):
    
    sheet.cell(row = 1, column = PLOT_SPECIAL_COLUMNS['Rs']).value = 'Rs-Linear-yData'
    
    for idx,dataPair in enumerate(cellDataObject.data,start=2):
        sheet.cell(row=idx, column = PLOT_SPECIAL_COLUMNS['Rs']).value = '=1 / H2 * A{} - D2 / H2'.format(idx + 2)
    
    dataX = Reference(sheet, min_col = 1, min_row = 4, max_row = len(cellDataObject.data[:,0]) + 3)
    dataY = Reference(sheet, min_col = PLOT_SPECIAL_COLUMNS['Rs'], min_row = 1, max_row = len(cellDataObject.data[:,0]) + 1)
    
    series = Series(dataY, xvalues = dataX, title_from_data=True)
    
    return series


